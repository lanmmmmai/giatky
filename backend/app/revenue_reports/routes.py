from fastapi import APIRouter, Depends, HTTPException, status, Query
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, date, timedelta, timezone
import uuid
import logging
from io import BytesIO
from fastapi.responses import StreamingResponse

from app.common.dependencies import get_current_user
from app.database import supabase

logger = logging.getLogger("app.revenue_reports")
router = APIRouter(prefix="/revenue-reports", tags=["Revenue Reports"])

# Pydantic schemas
class ReportCreate(BaseModel):
    report_date: date
    branch_id: str
    opening_cash: int = Field(..., ge=0)
    expense_amount: int = Field(0, ge=0)
    expense_description: Optional[str] = ""
    note: Optional[str] = ""

class ReportManualFieldsUpdate(BaseModel):
    opening_cash: int = Field(..., ge=0)
    expense_amount: int = Field(0, ge=0)
    expense_description: Optional[str] = ""
    note: Optional[str] = ""

class StatusChangeRequest(BaseModel):
    reject_reason: Optional[str] = None

# Helper functions

def to_number(value, default: int = 0) -> int:
    if value is None:
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default


def parse_db_date(value) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if not value:
        raise ValueError("Missing date value")
    return datetime.fromisoformat(str(value).replace("Z", "+00:00")).date()


def validate_branch_id(branch_id: str) -> None:
    try:
        uuid.UUID(str(branch_id))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="branch_id phải là UUID hợp lệ.")


def ensure_branch_access(branch_id: str, current_user: dict) -> dict:
    validate_branch_id(branch_id)
    branch_res = supabase.table("branches").select("*").eq("id", branch_id).limit(1).execute()
    if not branch_res.data:
        raise HTTPException(status_code=404, detail="Không tìm thấy cơ sở chi nhánh.")

    if current_user["role"] == "admin":
        return branch_res.data[0]

    if current_user["role"] == "manager":
        branch = branch_res.data[0]
        if branch.get("manager_id") == current_user["id"] or branch.get("created_by_admin_id") == current_user["id"]:
            return branch
        raise HTTPException(status_code=403, detail="Bạn không quản lý cơ sở này.")

    if current_user["role"] == "staff":
        allowed_branch_ids = {current_user.get("branch_id")}
        try:
            ub_res = supabase.table("user_branches").select("branch_id").eq("user_id", current_user["id"]).execute()
            allowed_branch_ids.update(item["branch_id"] for item in (ub_res.data or []) if item.get("branch_id"))
        except Exception as ub_err:
            logger.warning(
                "Failed to verify staff user_branches for revenue report access",
                extra={"user_id": current_user.get("id"), "branch_id": branch_id, "error": str(ub_err)},
            )
        if branch_id in allowed_branch_ids:
            return branch_res.data[0]
        raise HTTPException(status_code=403, detail="Bạn chỉ có thể xem báo cáo thuộc cơ sở của mình.")

    raise HTTPException(status_code=403, detail="Bạn không có quyền xem báo cáo doanh thu.")

def get_auto_calculations(branch_id: str, day_date: date):
    """Calculates order metrics and debt collections directly from the orders and debt_payments tables."""
    try:
        # Construct UTC range matching Vietnam local day (+07:00)
        vn_tz = timezone(timedelta(hours=7))
        start_local = datetime.combine(day_date, datetime.min.time()).replace(tzinfo=vn_tz)
        end_local = datetime.combine(day_date + timedelta(days=1), datetime.min.time()).replace(tzinfo=vn_tz)
        
        start_dt = start_local.astimezone(timezone.utc).isoformat()
        end_dt = end_local.astimezone(timezone.utc).isoformat()

        # Query orders created during this day range (excluding cancelled)
        orders_res = supabase.table("orders") \
            .select("total_amount, paid_amount, payment_method, payment_status, status") \
            .eq("branch_id", branch_id) \
            .neq("status", "cancelled") \
            .gte("created_at", start_dt) \
            .lt("created_at", end_dt) \
            .execute()
        orders = orders_res.data or []

        order_invoice_count = len(orders)
        order_bank_transfer = sum(to_number(o.get("paid_amount")) for o in orders if o.get("payment_method") == "bank_transfer")
        order_cash = sum(to_number(o.get("paid_amount")) for o in orders if o.get("payment_method") == "cash")
        order_debt = sum(to_number(o.get("total_amount")) - to_number(o.get("paid_amount")) for o in orders if o.get("payment_status") in ["unpaid", "partial"])

        # Query debt payments recorded on this date
        payments_res = supabase.table("debt_payments") \
            .select("amount, payment_method, order_id") \
            .eq("branch_id", branch_id) \
            .eq("payment_date", day_date.isoformat()) \
            .execute()
        payments = payments_res.data or []

        debt_invoice_count = len(set(p.get("order_id") for p in payments if p.get("order_id")))
        debt_bank_transfer = sum(to_number(p.get("amount")) for p in payments if p.get("payment_method") == "bank_transfer")
        debt_cash = sum(to_number(p.get("amount")) for p in payments if p.get("payment_method") == "cash")

        return {
            "order_invoice_count": order_invoice_count,
            "order_bank_transfer": order_bank_transfer,
            "order_cash": order_cash,
            "order_debt": order_debt,
            "daily_revenue": order_bank_transfer + order_cash + order_debt,
            "debt_invoice_count": debt_invoice_count,
            "debt_bank_transfer": debt_bank_transfer,
            "debt_cash": debt_cash,
            "debt_collection_total": debt_bank_transfer + debt_cash
        }
    except Exception as e:
        logger.exception(f"Error calculating automatic values for {day_date}: {str(e)}")
        return {
            "order_invoice_count": 0,
            "order_bank_transfer": 0,
            "order_cash": 0,
            "order_debt": 0,
            "daily_revenue": 0,
            "debt_invoice_count": 0,
            "debt_bank_transfer": 0,
            "debt_cash": 0,
            "debt_collection_total": 0
        }

# Endpoints

@router.get("/monthly")
def get_monthly_reports(
    branch_id: str,
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    current_user: dict = Depends(get_current_user)
):
    try:
        ensure_branch_access(branch_id, current_user)

        import calendar
        _, num_days = calendar.monthrange(year, month)

        # Get saved reports from database
        reports_res = supabase.table("revenue_reports") \
            .select("*") \
            .eq("branch_id", branch_id) \
            .eq("month", month) \
            .eq("year", year) \
            .execute()
        
        saved_reports = {}
        for r in (reports_res.data or []):
            d = parse_db_date(r["report_date"]).day
            saved_reports[d] = r

        # Optimize: Query all orders for the month in one batch
        start_date = date(year, month, 1)
        end_date_exclusive = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
        
        vn_tz = timezone(timedelta(hours=7))
        start_local = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=vn_tz)
        end_local = datetime.combine(end_date_exclusive, datetime.min.time()).replace(tzinfo=vn_tz)
        
        start_dt = start_local.astimezone(timezone.utc).isoformat()
        end_dt = end_local.astimezone(timezone.utc).isoformat()

        orders_res = supabase.table("orders") \
            .select("total_amount, paid_amount, payment_method, payment_status, created_at") \
            .eq("branch_id", branch_id) \
            .neq("status", "cancelled") \
            .gte("created_at", start_dt) \
            .lt("created_at", end_dt) \
            .execute()
        all_orders = orders_res.data or []

        # Optimize: Query all debt payments for the month in one batch
        try:
            payments_res = supabase.table("debt_payments") \
                .select("amount, payment_method, order_id, payment_date") \
                .eq("branch_id", branch_id) \
                .gte("payment_date", start_date.isoformat()) \
                .lt("payment_date", end_date_exclusive.isoformat()) \
                .execute()
            all_payments = payments_res.data or []
        except Exception as payments_err:
            logger.warning(
                "Debt payments query failed while building monthly revenue report",
                extra={
                    "branch_id": branch_id,
                    "month": month,
                    "year": year,
                    "user_id": current_user.get("id"),
                    "role": current_user.get("role"),
                    "error": str(payments_err),
                },
            )
            all_payments = []

        # Group orders by day in local +07:00 time
        orders_by_day = {}
        for o in all_orders:
            dt = datetime.fromisoformat(o["created_at"].replace("Z", "+00:00")).astimezone(vn_tz)
            day = dt.date().day
            if day not in orders_by_day:
                orders_by_day[day] = []
            orders_by_day[day].append(o)

        # Group payments by day
        payments_by_day = {}
        for p in all_payments:
            p_date = parse_db_date(p["payment_date"])
            day = p_date.day
            if day not in payments_by_day:
                payments_by_day[day] = []
            payments_by_day[day].append(p)

        # Build day-by-day reports array
        results = []
        previous_day_closing_cash = 0
        cumulative_revenue = 0

        # Try to pull the previous day's closing cash carry over (end of previous month)
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        _, prev_num_days = calendar.monthrange(prev_year, prev_month)
        prev_date = date(prev_year, prev_month, prev_num_days)
        prev_report_res = supabase.table("revenue_reports").select("closing_cash").eq("branch_id", branch_id).eq("report_date", prev_date.isoformat()).execute()
        if prev_report_res.data:
            previous_day_closing_cash = to_number(prev_report_res.data[0].get("closing_cash"))

        for day in range(1, num_days + 1):
            day_date = date(year, month, day)
            saved = saved_reports.get(day)
            
            day_orders = orders_by_day.get(day, [])
            day_payments = payments_by_day.get(day, [])

            # Automatic order metrics calculations
            order_invoice_count = len(day_orders)
            order_bank_transfer = sum(to_number(o.get("paid_amount")) for o in day_orders if o.get("payment_method") == "bank_transfer")
            order_cash = sum(to_number(o.get("paid_amount")) for o in day_orders if o.get("payment_method") == "cash")
            order_debt = sum(to_number(o.get("total_amount")) - to_number(o.get("paid_amount")) for o in day_orders if o.get("payment_status") in ["unpaid", "partial"])
            daily_revenue = order_bank_transfer + order_cash + order_debt
            
            # Automatic debt payments calculations
            debt_invoice_count = len(set(p.get("order_id") for p in day_payments if p.get("order_id")))
            debt_bank_transfer = sum(to_number(p.get("amount")) for p in day_payments if p.get("payment_method") == "bank_transfer")
            debt_cash = sum(to_number(p.get("amount")) for p in day_payments if p.get("payment_method") == "cash")
            debt_collection_total = debt_bank_transfer + debt_cash

            cumulative_revenue += daily_revenue

            # Resolve manual parameters and saved state
            if saved:
                report_id = saved["id"]
                if day == 1:
                    opening_cash = to_number(saved.get("opening_cash"))
                else:
                    opening_cash = previous_day_closing_cash
                expense_amount = to_number(saved.get("total_expense"))
                
                note_field = saved["note"] or ""
                if "|||" in note_field:
                    parts = note_field.split("|||", 1)
                    expense_description = parts[0]
                    note = parts[1]
                else:
                    expense_description = ""
                    note = note_field
                    
                status = saved["status"]
                approved_by = saved["approved_by"]
                approved_at = saved["approved_at"]
                reject_reason = saved["reject_reason"]
            else:
                report_id = None
                opening_cash = previous_day_closing_cash
                expense_amount = 0
                expense_description = ""
                note = ""
                status = "draft"
                approved_by = None
                approved_at = None
                reject_reason = None

            # Enforce mathematical closing balance
            closing_cash = opening_cash + order_cash + debt_cash - expense_amount
            previous_day_closing_cash = closing_cash

            results.append({
                "id": report_id,
                "report_date": day_date.isoformat(),
                "month": month,
                "year": year,
                "branch_id": branch_id,
                "opening_cash": opening_cash,
                "daily_revenue": daily_revenue,
                "cumulative_revenue": cumulative_revenue,
                "order_invoice_count": order_invoice_count,
                "order_bank_transfer": order_bank_transfer,
                "order_cash": order_cash,
                "order_debt": order_debt,
                "debt_collection_total": debt_collection_total,
                "debt_invoice_count": debt_invoice_count,
                "debt_bank_transfer": debt_bank_transfer,
                "debt_cash": debt_cash,
                "expense_amount": expense_amount,
                "expense_description": expense_description,
                "closing_cash": closing_cash,
                "note": note,
                "status": status,
                "approved_by": approved_by,
                "approved_at": approved_at,
                "reject_reason": reject_reason
            })
            
        return results
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(
            "GET /revenue-reports/monthly failed",
            extra={
                "branch_id": branch_id,
                "month": month,
                "year": year,
                "user_id": current_user.get("id"),
                "role": current_user.get("role"),
                "error": str(e),
            },
        )
        raise HTTPException(status_code=500, detail=f"Không thể tải danh sách báo cáo: {str(e)}")

@router.post("")
def create_report(payload: ReportCreate, current_user: dict = Depends(get_current_user)):
    try:
        # Verify unique index
        dup_res = supabase.table("revenue_reports").select("id").eq("report_date", payload.report_date.isoformat()).eq("branch_id", payload.branch_id).execute()
        if dup_res.data:
            raise HTTPException(status_code=400, detail=f"Báo cáo ngày {payload.report_date} cho cơ sở này đã tồn tại.")

        # Compute auto calculations for this date range
        auto = get_auto_calculations(payload.branch_id, payload.report_date)

        # Calculate cumulative revenue from beginning of the month
        reports_res = supabase.table("revenue_reports") \
            .select("daily_revenue") \
            .eq("branch_id", payload.branch_id) \
            .eq("month", payload.report_date.month) \
            .eq("year", payload.report_date.year) \
            .lt("report_date", payload.report_date.isoformat()) \
            .execute()
        
        cumulative_before = sum(r["daily_revenue"] for r in (reports_res.data or []))
        cumulative_revenue = cumulative_before + auto["daily_revenue"]

        # Delimit the text description and note parameters
        note_field = f"{payload.expense_description or ''}|||{payload.note or ''}"

        # Formula closing balance
        closing_cash = payload.opening_cash + auto["order_cash"] + auto["debt_cash"] - payload.expense_amount

        # Write data payload mapping to existing columns
        rep_data = {
            "report_date": payload.report_date.isoformat(),
            "month": payload.report_date.month,
            "year": payload.report_date.year,
            "branch_id": payload.branch_id,
            "staff_id": current_user["id"],
            "opening_cash": payload.opening_cash,
            "daily_revenue": auto["daily_revenue"],
            "cumulative_revenue_before": cumulative_revenue,
            "order_invoice_count": auto["order_invoice_count"],
            "order_bank_transfer": auto["order_bank_transfer"],
            "order_cash": auto["order_cash"],
            "order_debt": auto["order_debt"],
            "debt_collection_total": auto["debt_collection_total"],
            "debt_invoice_count": auto["debt_invoice_count"],
            "debt_bank_transfer": auto["debt_bank_transfer"],
            "debt_cash": auto["debt_cash"],
            "total_expense": payload.expense_amount,
            "note": note_field,
            "closing_cash": closing_cash,
            "status": "draft"
        }

        ins_res = supabase.table("revenue_reports").insert(rep_data).execute()
        if not ins_res.data:
            raise HTTPException(status_code=500, detail="Không thể lưu báo cáo.")

        return {"success": True, "data": ins_res.data[0]}
    except HTTPException:
        raise
    except Exception as e:
        print("POST /revenue-reports failed:", repr(e))
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/{id}/manual-fields")
def update_manual_fields(id: str, payload: ReportManualFieldsUpdate, current_user: dict = Depends(get_current_user)):
    try:
        rep_res = supabase.table("revenue_reports").select("*").eq("id", id).execute()
        if not rep_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo.")
        report = rep_res.data[0]

        if report["status"] == "approved" and current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Không thể sửa báo cáo đã được duyệt.")

        # Compute auto calculations for this date range to keep it synchronized
        day_date = datetime.strptime(report["report_date"], "%Y-%m-%d").date()
        auto = get_auto_calculations(report["branch_id"], day_date)

        # Delimit the text description and note parameters
        note_field = f"{payload.expense_description or ''}|||{payload.note or ''}"

        # Formula closing balance
        closing_cash = payload.opening_cash + auto["order_cash"] + auto["debt_cash"] - payload.expense_amount

        # Write data payload mapping to existing columns
        update_data = {
            "opening_cash": payload.opening_cash,
            "total_expense": payload.expense_amount,
            "note": note_field,
            "closing_cash": closing_cash,
            "daily_revenue": auto["daily_revenue"],
            "order_invoice_count": auto["order_invoice_count"],
            "order_bank_transfer": auto["order_bank_transfer"],
            "order_cash": auto["order_cash"],
            "order_debt": auto["order_debt"],
            "debt_collection_total": auto["debt_collection_total"],
            "debt_invoice_count": auto["debt_invoice_count"],
            "debt_bank_transfer": auto["debt_bank_transfer"],
            "debt_cash": auto["debt_cash"],
            "updated_at": datetime.utcnow().isoformat()
        }

        upd_res = supabase.table("revenue_reports").update(update_data).eq("id", id).execute()
        return {"success": True, "data": upd_res.data[0]}
    except HTTPException:
        raise
    except Exception as e:
        print("PUT /revenue-reports/{id}/manual-fields failed:", repr(e))
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{id}")
def delete_report(id: str, current_user: dict = Depends(get_current_user)):
    try:
        rep_res = supabase.table("revenue_reports").select("*").eq("id", id).execute()
        if not rep_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo.")
        report = rep_res.data[0]

        if current_user["role"] == "staff":
            if report["status"] == "approved":
                raise HTTPException(status_code=403, detail="Không thể xóa báo cáo đã được duyệt.")
            if report["branch_id"] != current_user.get("branch_id"):
                raise HTTPException(status_code=403, detail="Không thể xóa báo cáo của chi nhánh khác.")
        elif current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Chỉ có Quản trị viên mới có quyền xóa báo cáo.")

        supabase.table("revenue_reports").delete().eq("id", id).execute()
        return {"success": True, "message": "Xóa báo cáo doanh thu thành công."}
    except HTTPException:
        raise
    except Exception as e:
        print("DELETE /revenue-reports/{id} failed:", repr(e))
        raise HTTPException(status_code=500, detail=f"Không thể xóa báo cáo: {str(e)}")

# Workflow State Changes

@router.patch("/{id}/submit")
def submit_report(id: str, current_user: dict = Depends(get_current_user)):
    try:
        rep_res = supabase.table("revenue_reports").select("*").eq("id", id).execute()
        if not rep_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo.")
        report = rep_res.data[0]

        if current_user["role"] == "staff" and report["branch_id"] != current_user.get("branch_id"):
            raise HTTPException(status_code=403, detail="Bạn không thể nộp báo cáo của chi nhánh khác.")

        upd = supabase.table("revenue_reports").update({
            "status": "submitted",
            "updated_at": datetime.utcnow().isoformat()
        }).eq("id", id).execute()
        
        return upd.data[0]
    except HTTPException:
        raise
    except Exception as e:
        print("PATCH /revenue-reports/{id}/submit failed:", repr(e))
        raise HTTPException(status_code=500, detail=f"Không thể nộp báo cáo: {str(e)}")

@router.patch("/{id}/approve")
def approve_report(id: str, current_user: dict = Depends(get_current_user)):
    try:
        if current_user["role"] not in ["admin", "manager"]:
            raise HTTPException(status_code=403, detail="Chỉ Quản lý hoặc Quản trị viên mới được duyệt báo cáo.")

        rep_res = supabase.table("revenue_reports").select("*").eq("id", id).execute()
        if not rep_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo.")
        report = rep_res.data[0]

        if current_user["role"] == "manager":
            br_res = supabase.table("branches").select("id").eq("id", report["branch_id"]).eq("manager_id", current_user["id"]).execute()
            if not br_res.data:
                raise HTTPException(status_code=403, detail="Bạn chỉ có thể duyệt báo cáo thuộc chi nhánh mình quản lý.")

        upd = supabase.table("revenue_reports").update({
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_at": datetime.utcnow().isoformat(),
            "reject_reason": None,
            "updated_at": datetime.utcnow().isoformat()
        }).eq("id", id).execute()

        return upd.data[0]
    except HTTPException:
        raise
    except Exception as e:
        print("PATCH /revenue-reports/{id}/approve failed:", repr(e))
        raise HTTPException(status_code=500, detail=f"Không thể duyệt báo cáo: {str(e)}")

@router.patch("/{id}/reject")
def reject_report(id: str, payload: StatusChangeRequest, current_user: dict = Depends(get_current_user)):
    try:
        if current_user["role"] not in ["admin", "manager"]:
            raise HTTPException(status_code=403, detail="Chỉ Quản lý hoặc Quản trị viên mới có quyền trả lại báo cáo.")

        rep_res = supabase.table("revenue_reports").select("*").eq("id", id).execute()
        if not rep_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo.")
        report = rep_res.data[0]

        if current_user["role"] == "manager":
            br_res = supabase.table("branches").select("id").eq("id", report["branch_id"]).eq("manager_id", current_user["id"]).execute()
            if not br_res.data:
                raise HTTPException(status_code=403, detail="Bạn chỉ có quyền chỉnh sửa/trả lại báo cáo thuộc chi nhánh của mình.")

        upd = supabase.table("revenue_reports").update({
            "status": "rejected",
            "reject_reason": payload.reject_reason or "Không được duyệt bởi quản lý.",
            "updated_at": datetime.utcnow().isoformat()
        }).eq("id", id).execute()

        return upd.data[0]
    except HTTPException:
        raise
    except Exception as e:
        print("PATCH /revenue-reports/{id}/reject failed:", repr(e))
        raise HTTPException(status_code=500, detail=f"Không thể trả lại báo cáo: {str(e)}")

# Excel Export

@router.get("/export-excel")
def export_excel(
    branch_id: str,
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    current_user: dict = Depends(get_current_user)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import urllib.parse

        # Call monthly list reports to build data structure
        reports = get_monthly_reports(branch_id, month, year, current_user)

        branch_res = supabase.table("branches").select("name").eq("id", branch_id).execute()
        if not branch_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy cơ sở chi nhánh.")
        branch_name = branch_res.data[0]["name"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"T{month}{year}"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
        header_font = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
        data_font = Font(name="Calibri", size=10)
        bold_font = Font(name="Calibri", size=10, bold=True)
        
        header_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # soft green
        highlight_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # soft yellow
        summary_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        double_bottom_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='double', color='1E3A8A')
        )

        # Title
        ws.merge_cells("A1:P1")
        ws["A1"] = f"BÁO CÁO DOANH THU THÁNG {month}/{year} - CƠ SỞ {branch_name.upper()}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Subtitle
        ws.merge_cells("A2:P2")
        ws["A2"] = "Cửa hàng: Giặt Ký | Sạch Thơm Tin Tưởng"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="475569")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 10

        # Group headers
        headers = [
            ("STT", "A4:A5"),
            ("Ngày", "B4:B5"),
            ("Đầu kỳ", "C4:C5"),
            ("Doanh thu ngày báo cáo", "D4:D5"),
            ("Nhận đơn: xHD", "E4:H4"),
            ("Thu nợ: xHD", "I4:L4"),
            ("Phát sinh", "M4:N4"),
            ("Tổng tiền tại CH", "O4:O5"),
            ("Ghi chú", "P4:P5")
        ]
        
        sub_headers = {
            "E5": "Số HD", "F5": "CK", "G5": "TM", "H5": "Nợ",
            "I5": "Tổng thu nợ", "J5": "Số HD", "K5": "CK", "L5": "TM",
            "M5": "Số tiền", "N5": "Diễn giải"
        }

        # Apply header formatting
        for h, cell_range in headers:
            ws.merge_cells(cell_range)
            first_cell = cell_range.split(":")[0]
            ws[first_cell] = h
            
        for coord, text in sub_headers.items():
            ws[coord] = text

        for row in range(4, 6):
            ws.row_dimensions[row].height = 25
            for col in range(1, 17):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

        # Insert data
        row_idx = 6
        for idx, rep in enumerate(reports):
            ws.row_dimensions[row_idx].height = 22
            
            day = idx + 1
            ws.cell(row=row_idx, column=1, value=day).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=f"{day:02d}/{month:02d}/{year}").alignment = Alignment(horizontal="center")

            ws.cell(row=row_idx, column=3, value=rep["opening_cash"])
            ws.cell(row=row_idx, column=4, value=rep["daily_revenue"])
            
            # Nhận đơn
            ws.cell(row=row_idx, column=5, value=rep["order_invoice_count"])
            ws.cell(row=row_idx, column=6, value=rep["order_bank_transfer"])
            ws.cell(row=row_idx, column=7, value=rep["order_cash"])
            ws.cell(row=row_idx, column=8, value=rep["order_debt"])
            
            # Thu nợ
            ws.cell(row=row_idx, column=9, value=rep["debt_collection_total"])
            ws.cell(row=row_idx, column=10, value=rep["debt_invoice_count"])
            ws.cell(row=row_idx, column=11, value=rep["debt_bank_transfer"])
            ws.cell(row=row_idx, column=12, value=rep["debt_cash"])
            
            # Chi phí
            ws.cell(row=row_idx, column=13, value=rep["expense_amount"])
            ws.cell(row=row_idx, column=14, value=rep["expense_description"])
            
            # Kết quả
            ws.cell(row=row_idx, column=15, value=rep["closing_cash"])
            ws.cell(row=row_idx, column=16, value=rep["note"])

            for col in range(1, 17):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Highlight cumulative/daily/debt/closing fields in soft yellow
                if col in [4, 9, 15]:
                    cell.fill = highlight_fill
                
                if col in [3, 4, 6, 7, 8, 9, 11, 12, 13, 15]:
                    cell.number_format = '#,##0" đ"'
                    cell.alignment = Alignment(horizontal="right")
                elif col in [1, 2, 5, 10]:
                    cell.alignment = Alignment(horizontal="center")
                    
            row_idx += 1

        # Totals bottom row
        ws.row_dimensions[row_idx].height = 26
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        ws.cell(row=row_idx, column=1, value="TỔNG CỘNG").font = bold_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")

        formula_cols = {
            3: f"=SUM(C6:C{row_idx-1})",
            4: f"=SUM(D6:D{row_idx-1})",
            5: f"=SUM(E6:E{row_idx-1})",
            6: f"=SUM(F6:F{row_idx-1})",
            7: f"=SUM(G6:G{row_idx-1})",
            8: f"=SUM(H6:H{row_idx-1})",
            9: f"=SUM(I6:I{row_idx-1})",
            10: f"=SUM(J6:J{row_idx-1})",
            11: f"=SUM(K6:K{row_idx-1})",
            12: f"=SUM(L6:L{row_idx-1})",
            13: f"=SUM(M6:M{row_idx-1})",
            15: f"=SUM(O6:O{row_idx-1})"
        }

        for col, formula in formula_cols.items():
            cell = ws.cell(row=row_idx, column=col, value=formula)
            cell.font = bold_font
            cell.number_format = '#,##0" đ"'
            cell.alignment = Alignment(horizontal="right", vertical="center")

        for col in range(1, 17):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = summary_fill
            cell.border = double_bottom_border

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and ' đ' in cell.number_format:
                    val += " đ"
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        filename = f"bao_cao_doanh_thu_T{month}_{year}_{branch_name.replace(' ', '_')}.xlsx"
        safe_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=utf-8''{safe_filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        print("GET /revenue-reports/export-excel failed:", repr(e))
        raise HTTPException(status_code=500, detail=f"Không thể xuất báo cáo Excel: {str(e)}")
